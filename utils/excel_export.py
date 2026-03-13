"""
Excel Export Functionality for Google Maps Scraper
"""
import os
from datetime import datetime
from typing import List, Optional
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.models import Isletme

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Excel export sınıfı"""
    
    def __init__(self, output_dir: str = "exports"):
        """
        Excel exporter başlat
        
        Args:
            output_dir: Export klasörü
        """
        self.output_dir = output_dir
        self._ensure_output_dir()
        
        # Stil tanımlamaları
        self.header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.data_font = Font(name='Calibri', size=10)
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Border
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def _ensure_output_dir(self):
        """Export klasörünü oluştur"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def export_businesses(self, businesses: List[Isletme], filename: str = None) -> str:
        """
        İşletmeleri Excel'e streaming export et (memory-efficient)
        
        Args:
            businesses: İşletme listesi
            filename: Dosya adı (opsiyonel)
            
        Returns:
            str: Oluşturulan dosya yolu
        """
        if not businesses:
            raise ValueError("Export edilecek işletme bulunamadı")
        
        # Memory usage kontrolü
        total_businesses = len(businesses)
        if total_businesses > 10000:
            logger.warning(f"Büyük veri seti tespit edildi: {total_businesses} kayıt. Streaming export kullanılıyor.")
        
        # Dosya adı oluştur
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"google_maps_businesses_{timestamp}"
        
        # .xlsx uzantısı ekle
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Workbook oluştur
        wb = None
        ws = None
        
        try:
            wb = Workbook(write_only=True)  # Write-only mode for memory efficiency
            ws = wb.active
            ws.title = "İşletmeler"
            
            # Header'ları yaz
            headers = [
                'Sıra No', 'İşletme Adı', 'Kategori', 'Telefon', 'Adres',
                'Website', 'Çalışma Saatleri', 'Puan', 'Yorum Sayısı',
                'Yoğunluk Bilgisi', 'İl', 'İlçe', 'Durum', 'Notlar',
                'Konum Linki', 'Resim URL', 'Google ID', 'Eklenme Tarihi'
            ]
            
            # Header satırını yaz
            header_row = []
            for header in headers:
                cell = openpyxl.cell.WriteOnlyCell(ws, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
                header_row.append(cell)
            
            ws.append(header_row)
            
            # Veri satırlarını batch'ler halinde yaz (memory efficient)
            batch_size = 1000
            for batch_start in range(0, total_businesses, batch_size):
                batch_end = min(batch_start + batch_size, total_businesses)
                batch = businesses[batch_start:batch_end]
                
                for i, business in enumerate(batch):
                    row_num = batch_start + i + 2  # +2 because of header and 1-based indexing
                    
                    # Veri hazırla
                    data = [
                        row_num - 1,  # Sıra no
                        business.isim or '',
                        business.kategori or '',
                        business.telefon or '',
                        business.adres or '',
                        business.website or '',
                        business.calisma_saatleri or '',
                        business.puan or '',
                        business.yorum_sayisi or 0,
                        business.yogunluk_bilgisi or '',
                        business.il or '',
                        business.ilce or '',
                        'Müşteri' if business.durum == 1 else 'Potansiyel',
                        business.notlar or '',
                        business.konum_linki or '',
                        business.resim_url or '',
                        business.google_id or '',
                        business.eklenme_tarihi.strftime('%Y-%m-%d %H:%M:%S') if business.eklenme_tarihi else ''
                    ]
                    
                    # Satırı yaz
                    data_row = []
                    for col, value in enumerate(data, 1):
                        cell = openpyxl.cell.WriteOnlyCell(ws, value=value)
                        cell.font = self.data_font
                        cell.border = self.border
                        
                        # Hizalama
                        if col in [1, 8, 9]:  # Sıra no, puan, yorum sayısı
                            cell.alignment = self.center_alignment
                        else:
                            cell.alignment = self.left_alignment
                        
                        data_row.append(cell)
                    
                    ws.append(data_row)
                
                # Batch tamamlandığında memory'yi temizle
                if batch_end < total_businesses:
                    logger.debug(f"Batch tamamlandı: {batch_end}/{total_businesses}")
            
            # Sütun genişliklerini ayarla (write-only mode'da sınırlı)
            # Not: write-only mode'da column dimensions ayarlanamaz
            
            # Dosyayı kaydet
            wb.save(filepath)
            logger.info(f"Excel dosyası oluşturuldu: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Excel export hatası: {e}")
            raise
        finally:
            # Memory cleanup
            if wb:
                try:
                    wb.close()
                except:
                    pass
            wb = None
            ws = None
        
        # Sütun genişliklerini ayarla
        column_widths = {
            'A': 8,   # Sıra No
            'B': 30,  # İşletme Adı
            'C': 20,  # Kategori
            'D': 15,  # Telefon
            'E': 40,  # Adres
            'F': 25,  # Website
            'G': 25,  # Çalışma Saatleri
            'H': 8,   # Puan
            'I': 12,  # Yorum Sayısı
            'J': 20,  # Yoğunluk Bilgisi
            'K': 12,  # İl
            'L': 15,  # İlçe
            'M': 12,  # Durum
            'N': 30,  # Notlar
            'O': 15,  # Konum Linki
            'P': 15,  # Resim URL
            'Q': 20,  # Google ID
            'R': 18   # Eklenme Tarihi
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze panes (header'ı sabitle)
        ws.freeze_panes = 'A2'
        
        # Auto filter ekle
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        
        # İstatistik sayfası ekle
        self._add_statistics_sheet(wb, businesses)
        
        # Dosyayı kaydet
        wb.save(filepath)
        logger.info(f"Excel dosyası oluşturuldu: {filepath}")
        
        return filepath
    
    def _add_statistics_sheet(self, wb: Workbook, businesses: List[Isletme]):
        """İstatistik sayfası ekle"""
        ws_stats = wb.create_sheet("İstatistikler")
        
        # İstatistikleri hesapla
        total_count = len(businesses)
        customer_count = sum(1 for b in businesses if b.durum == 1)
        potential_count = total_count - customer_count
        phone_count = sum(1 for b in businesses if b.telefon)
        website_count = sum(1 for b in businesses if b.website)
        rated_count = sum(1 for b in businesses if b.puan)
        
        # İl bazında dağılım
        il_stats = {}
        for business in businesses:
            il = business.il or 'Bilinmeyen'
            il_stats[il] = il_stats.get(il, 0) + 1
        
        # Kategori bazında dağılım
        kategori_stats = {}
        for business in businesses:
            kategori = business.kategori or 'Bilinmeyen'
            kategori_stats[kategori] = kategori_stats.get(kategori, 0) + 1
        
        # İstatistik verilerini yaz
        stats_data = [
            ['GENEL İSTATİSTİKLER', ''],
            ['Toplam İşletme Sayısı', total_count],
            ['Müşteri Sayısı', customer_count],
            ['Potansiyel Müşteri Sayısı', potential_count],
            ['Telefonu Olan İşletme', phone_count],
            ['Websitesi Olan İşletme', website_count],
            ['Puanı Olan İşletme', rated_count],
            ['', ''],
            ['İL BAZINDA DAĞILIM', '']
        ]
        
        # İl istatistiklerini ekle
        for il, count in sorted(il_stats.items()):
            stats_data.append([il, count])
        
        stats_data.extend([['', ''], ['KATEGORİ BAZINDA DAĞILIM', '']])
        
        # Kategori istatistiklerini ekle (en çok olan 10 tanesi)
        sorted_categories = sorted(kategori_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        for kategori, count in sorted_categories:
            stats_data.append([kategori, count])
        
        # Verileri sayfaya yaz
        for row, (label, value) in enumerate(stats_data, 1):
            # Label
            cell_label = ws_stats.cell(row=row, column=1, value=label)
            if label in ['GENEL İSTATİSTİKLER', 'İL BAZINDA DAĞILIM', 'KATEGORİ BAZINDA DAĞILIM']:
                cell_label.font = Font(name='Calibri', size=12, bold=True)
                cell_label.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            else:
                cell_label.font = Font(name='Calibri', size=10)
            
            # Value
            if value != '':
                cell_value = ws_stats.cell(row=row, column=2, value=value)
                cell_value.font = Font(name='Calibri', size=10)
                cell_value.alignment = self.center_alignment
        
        # Sütun genişliklerini ayarla
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 15
        
        # Export tarihi ekle
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_stats.cell(row=len(stats_data) + 3, column=1, value=f"Export Tarihi: {export_date}")
    
    def export_customers(self, customers_data: List[dict], filename: str = None) -> str:
        """
        Müşterileri Excel'e export et - Resource-safe
        
        Args:
            customers_data: Müşteri verileri
            filename: Dosya adı
            
        Returns:
            str: Oluşturulan dosya yolu
        """
        if not customers_data:
            raise ValueError("Export edilecek müşteri bulunamadı")
        
        # Dosya adı oluştur
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"musteriler_{timestamp}"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Workbook oluştur - context manager ile resource-safe
        wb = None
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Müşteriler"
            
            # Header'ları yaz
            headers = [
                'Sıra No', 'İşletme Adı', 'Kategori', 'Telefon', 'Adres',
                'İl', 'İlçe', 'Paket', 'Ödeme Durumu', 'İletişim Tarihi',
                'Müşteri Notları', 'Kayıt Tarihi'
            ]
            
            # Header satırını yaz
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Veri satırlarını batch'ler halinde yaz (memory efficient)
            batch_size = 500
            total_customers = len(customers_data)
            
            for batch_start in range(0, total_customers, batch_size):
                batch_end = min(batch_start + batch_size, total_customers)
                batch = customers_data[batch_start:batch_end]
                
                for i, customer in enumerate(batch):
                    row = batch_start + i + 2  # +2 for header and 1-based indexing
                    
                    data = [
                        row - 1,  # Sıra no
                        customer.get('isim', ''),
                        customer.get('kategori', ''),
                        customer.get('telefon', ''),
                        customer.get('adres', ''),
                        customer.get('il', ''),
                        customer.get('ilce', ''),
                        customer.get('paket', ''),
                        customer.get('odeme_durumu', ''),
                        customer.get('iletisim_tarihi', ''),
                        customer.get('notlar', ''),
                        customer.get('kayit_tarihi', '')
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.data_font
                        cell.border = self.border
                        
                        if col == 1:  # Sıra no
                            cell.alignment = self.center_alignment
                        else:
                            cell.alignment = self.left_alignment
                
                # Batch progress log
                if batch_end < total_customers:
                    logger.debug(f"Customer export batch: {batch_end}/{total_customers}")
            
            # Sütun genişliklerini ayarla
            column_widths = {
                'A': 8,   # Sıra No
                'B': 30,  # İşletme Adı
                'C': 20,  # Kategori
                'D': 15,  # Telefon
                'E': 40,  # Adres
                'F': 12,  # İl
                'G': 15,  # İlçe
                'H': 20,  # Paket
                'I': 15,  # Ödeme Durumu
                'J': 15,  # İletişim Tarihi
                'K': 30,  # Müşteri Notları
                'L': 18   # Kayıt Tarihi
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Freeze panes ve auto filter
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
            
            # Dosyayı kaydet
            wb.save(filepath)
            logger.info(f"Müşteri Excel dosyası oluşturuldu: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Customer Excel export hatası: {e}")
            raise
        finally:
            # Resource cleanup
            if wb:
                try:
                    wb.close()
                except:
                    pass
            wb = None
    
    def get_export_directory(self) -> str:
        """Export klasörü yolunu al"""
        return os.path.abspath(self.output_dir)